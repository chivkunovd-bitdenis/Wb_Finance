from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO
from statistics import median, mean
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class ParseError(Exception):
    message: str


def parse_wb_competitor_excel(
    *,
    content: bytes,
    report_date: date,
    period: str,
    raw_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """
    Parse WB «Сравнение карточек» Excel into canonical payload for `import_competitor_report`.

    Semantics (лист «Показатели», колонки «Артикул WB …» — один из артикулов это наш товар).
    **Только эти подписи в первой колонке** (без альтернатив и «похожих» строк):
    - **Показы** → `traffic`: абсолют; по конкурентам — **среднее арифметическое** (нули среди конкурентов не берём).
    - **Конверсия в корзину, %** → `funnel_cart`: приводим всю строку к **процентным пунктам** (смешение п.п. и долей Excel как у CTR), затем по конкурентам — **медиана** (нули не берём).
    - **Конверсия в заказ, %** → `funnel_order`: то же.
    - **CTR** → `ctr`: как у конверсий — в БД **процентные пункты** (`unit` = «%»). Если в ячейке **доля** (строго между 0 и 1), умножаем на 100; иначе считаем, что уже п.п. По конкурентам — **медиана** (нули не берём).
    - **Количество отзывов** → `review_count` (целое; по конкурентам — медиана, для правил берётся **наше** значение).
    - **Рейтинг по отзывам** → `review_rating` (по конкурентам — медиана; для правил — **наше** значение).

    NOTE: Логистика и прочие затраты из финблока приложения сюда не входят — они берутся из `sku_daily` / аналитики, не из этого Excel.

    Contract output:
      {
        "report_date": <date>,
        "period": <period>,
        "source": "playwright",
        "raw_payload": {...},
        "items": [{"nm_id": int, "metric_code": str, "our_value": float|None, "competitor_median_value": float|None, "unit": str|None, "extra": dict|None}, ...]
      }

    NOTE: WB Excel structure may change. MVP parser is intentionally strict with a single supported sheet format.
    """
    if not content:
        raise ParseError("Empty excel payload")
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ParseError("Failed to read excel") from exc

    if not wb.worksheets:
        raise ParseError("Excel has no worksheets")

    def _norm(v: object) -> str:
        return str(v or "").strip().lower().replace(" ", "_")

    def _to_float_or_none(v: object) -> float | None:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(" ", "").replace(",", ".")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None

    # Format A (WB "Показатели" sheet): metrics in rows, nm_id in columns.
    if "Показатели" in wb.sheetnames:
        sh = wb["Показатели"]
        # openpyxl read_only may report max_column=1 for wide sheets; cap max_col explicitly.
        header = list(next(sh.iter_rows(min_row=2, max_row=2, max_col=300, values_only=True), []))
        if header and _norm(header[0]) in {"показатели", "показатель"}:
            # Collect nm_id columns for current period: "Артикул WB <id>" (exclude "Разница" and previous period).
            nm_cols: list[tuple[int, int]] = []
            for idx, h in enumerate(header):
                hs = str(h or "")
                if "(предыдущий период)" in hs:
                    continue
                if "Разница" in hs:
                    continue
                if "Артикул WB" in hs:
                    parts = hs.replace("Артикул WB", "").strip().split()
                    nm_raw = parts[0] if parts else ""
                    if nm_raw.isdigit():
                        nm_cols.append((int(nm_raw), idx))

            if not nm_cols:
                raise ParseError("Показатели: no nm_id columns found")

            def _read_metric_row(r_idx: int) -> list[Any]:
                return list(next(sh.iter_rows(min_row=r_idx, max_row=r_idx, max_col=300, values_only=True), []))

            # First label in column A -> row index (first occurrence wins).
            row_by_norm: dict[str, int] = {}
            for r_idx, r in enumerate(sh.iter_rows(min_row=3, max_row=200, values_only=True), start=3):
                name = _norm(r[0] if r else "")
                if name and name not in row_by_norm:
                    row_by_norm[name] = r_idx

            def _cells_for_nm(row_vals: list[Any]) -> dict[int, float | None]:
                out: dict[int, float | None] = {}
                for nm_id, c_idx in nm_cols:
                    out[nm_id] = _to_float_or_none(row_vals[c_idx] if c_idx < len(row_vals) else None)
                return out

            def _competitor_values_excluding(
                target_nm: int, values_by_nm: dict[int, float | None], *, exclude_zero: bool
            ) -> list[float]:
                out: list[float] = []
                for nm, v in values_by_nm.items():
                    if nm == target_nm or v is None:
                        continue
                    fv = float(v)
                    if exclude_zero and fv == 0.0:
                        continue
                    out.append(fv)
                return out

            def _median_excluding(target_nm: int, values_by_nm: dict[int, float | None]) -> float | None:
                """Медиана процентов по конкурентам; None и (опционально) 0 не участвуют."""
                others = _competitor_values_excluding(target_nm, values_by_nm, exclude_zero=True)
                return float(median(others)) if others else None

            def _mean_excluding(target_nm: int, values_by_nm: dict[int, float | None]) -> float | None:
                """Среднее по конкурентам для «Показы»; None и 0 не участвуют."""
                others = _competitor_values_excluding(target_nm, values_by_nm, exclude_zero=True)
                return float(mean(others)) if others else None

            def _ctr_cell_to_percent_points(v: float | None) -> float | None:
                """CTR в Excel: либо уже п.п. (3.1), либо доля (0.031) — тогда ×100 → п.п."""
                if v is None:
                    return None
                fv = float(v)
                if 0.0 < fv < 1.0:
                    return fv * 100.0
                return fv

            def _normalize_funnel_row_to_percent_points(by_nm: dict[int, float | None]) -> dict[int, float | None]:
                """
                WB часто отдаёт конверсии в одной строке в одной шкале, но встречается смесь:
                - числа 8, 12, 15 — уже процентные пункты;
                - доли 0.08–0.99 — как в Excel «Процент» (0.12 = 12%);
                - рядом с п.п. доля 1.0 часто означает отображаемые 100%, а не «1 п.п.».

                Медиану по конкурентам считаем **после** приведения всех ячеек строки к п.п.,
                иначе median(15, 0.12, 0.18) давала бы мусор вместо median(15, 12, 18).
                """
                vals = [float(v) for v in by_nm.values() if v is not None]
                if not vals:
                    return dict(by_nm)
                mx = max(vals)
                # Строго (0,1): доля Excel; ровно 1.0 без других долей — чаще «1 п.п.» в тестах/WB, не умножаем всю строку.
                any_open_fraction = any(0.0 < v < 1.0 for v in vals)
                mixed_scale = mx > 1.0 and any_open_fraction

                def _one(v: float | None) -> float | None:
                    if v is None:
                        return None
                    fv = float(v)
                    if mx <= 1.0 and any_open_fraction:
                        return fv * 100.0
                    if mixed_scale:
                        if 0.0 < fv < 1.0:
                            return fv * 100.0
                        if fv == 1.0:
                            return 100.0
                        return fv
                    return fv

                return {nm: _one(v) for nm, v in by_nm.items()}

            def _items_for_row(
                *,
                row_norm: str,
                metric_code: str,
                unit: str,
                aggregate: str,
            ) -> list[dict[str, Any]]:
                r_idx = row_by_norm.get(row_norm)
                if r_idx is None:
                    return []
                by_nm = _cells_for_nm(_read_metric_row(r_idx))
                if metric_code in {"funnel_cart", "funnel_order"}:
                    by_nm = _normalize_funnel_row_to_percent_points(by_nm)
                if metric_code == "ctr":
                    by_nm = {nm: _ctr_cell_to_percent_points(val) for nm, val in by_nm.items()}
                out: list[dict[str, Any]] = []
                for nm_id, _c in nm_cols:
                    our = by_nm.get(nm_id)
                    if our is None:
                        continue
                    if aggregate == "mean":
                        comp = _mean_excluding(nm_id, by_nm)
                    else:
                        comp = _median_excluding(nm_id, by_nm)
                    out.append(
                        {
                            "nm_id": nm_id,
                            "metric_code": metric_code,
                            "our_value": float(our),
                            "competitor_median_value": comp,
                            "unit": unit,
                            "extra": {"source": "excel_row", "competitor_aggregate": aggregate},
                        }
                    )
                return out

            # Ровно четыре обязательные строки WB (подпись в колонке A после нормализации).
            row_shows = _norm("Показы")
            row_cart = _norm("Конверсия в корзину, %")
            row_order = _norm("Конверсия в заказ, %")
            row_ctr = _norm("CTR")
            required = (
                (row_shows, "Показы"),
                (row_cart, "Конверсия в корзину, %"),
                (row_order, "Конверсия в заказ, %"),
                (row_ctr, "CTR"),
            )
            for key, human in required:
                if key not in row_by_norm:
                    raise ParseError(f"Показатели: нет строки «{human}» (ожидается точное название поля).")

            items: list[dict[str, Any]] = []
            items.extend(_items_for_row(row_norm=row_shows, metric_code="traffic", unit="шт", aggregate="mean"))
            items.extend(_items_for_row(row_norm=row_cart, metric_code="funnel_cart", unit="%", aggregate="median"))
            items.extend(_items_for_row(row_norm=row_order, metric_code="funnel_order", unit="%", aggregate="median"))
            items.extend(_items_for_row(row_norm=row_ctr, metric_code="ctr", unit="%", aggregate="median"))

            # Опционально (новые выгрузки WB): отзывы и рейтинг — для правила self_buyouts без ручного social.
            row_reviews = _norm("Количество отзывов")
            row_rating = _norm("Рейтинг по отзывам")
            if row_reviews in row_by_norm:
                items.extend(
                    _items_for_row(row_norm=row_reviews, metric_code="review_count", unit="шт", aggregate="median")
                )
            if row_rating in row_by_norm:
                items.extend(
                    _items_for_row(row_norm=row_rating, metric_code="review_rating", unit=None, aggregate="median")
                )

            if not items:
                raise ParseError("Показатели: no metrics parsed")

            return {
                "report_date": report_date,
                "period": period,
                "source": "playwright",
                "raw_payload": raw_payload,
                "items": items,
            }

    # Format B (legacy MVP): look for first sheet that has expected headers in rows.
    sheet = wb.worksheets[0]

    header_row: list[str] | None = None
    header_row_idx: int | None = None
    for i, r in enumerate(sheet.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        if not r:
            continue
        nn = [_norm(x) for x in r]
        if "nm_id" in nn or "артикул" in nn:
            header_row = nn
            header_row_idx = i
            break
    if header_row is None:
        raise ParseError("Excel header row not found")
    if header_row_idx is None:
        raise ParseError("Excel header row index not found")

    # Map columns
    col_index: dict[str, int] = {name: i for i, name in enumerate(header_row) if name}

    def _col(*names: str) -> int | None:
        for n in names:
            if n in col_index:
                return col_index[n]
        return None

    nm_col = _col("nm_id", "артикул")
    if nm_col is None:
        raise ParseError("nm_id column not found")

    metric_cols = {
        "ctr": (_col("ctr"), _col("ctr_median", "median_ctr", "ctr_медиана")),
        "traffic": (_col("traffic"), _col("traffic_median", "median_traffic", "traffic_медиана")),
        "funnel_cart": (_col("funnel_cart", "to_cart", "воронка_в_корзину"), _col("funnel_cart_median", "median_funnel_cart")),
        "funnel_order": (_col("funnel_order", "to_order", "воронка_в_заказ"), _col("funnel_order_median", "median_funnel_order")),
    }

    legacy_items: list[dict[str, Any]] = []

    # Data starts after the detected header row.
    for r in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not r:
            continue
        try:
            nm_raw = r[nm_col]
            if nm_raw in {None, ""}:
                continue
            nm_id = int(str(nm_raw).strip())
        except Exception:
            continue

        for code, (our_i, med_i) in metric_cols.items():
            if our_i is None and med_i is None:
                continue
            our_v = None
            med_v = None
            try:
                if our_i is not None:
                    ov = r[our_i]
                    our_v = float(ov) if ov not in {None, ""} else None
            except Exception:
                our_v = None
            try:
                if med_i is not None:
                    mv = r[med_i]
                    med_v = float(mv) if mv not in {None, ""} else None
            except Exception:
                med_v = None

            if our_v is None and med_v is None:
                continue
            legacy_items.append(
                {
                    "nm_id": nm_id,
                    "metric_code": code,
                    "our_value": our_v,
                    "competitor_median_value": med_v,
                    "unit": None,
                    "extra": None,
                }
            )

    if not legacy_items:
        raise ParseError("No metrics parsed from excel")

    return {
        "report_date": report_date,
        "period": period,
        "source": "playwright",
        "raw_payload": raw_payload,
        "items": legacy_items,
    }

