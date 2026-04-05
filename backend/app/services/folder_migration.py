from __future__ import annotations

import csv
import re
import secrets
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.models.raw_ads import RawAd
from app.models.raw_sales import RawSale
from app.models.user import User
from app.core.security import hash_password
from app.schemas.sync import FolderFileReport, FolderMigrationRequest, FolderMigrationResponse


@dataclass
class ParsedFile:
    path: Path
    user_email: str
    dataset: str | None = None


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    return float(str(value).replace(" ", "").replace(",", "."))


def _to_int(value: Any, default: int | None = None) -> int | None:
    if value is None or value == "":
        return default
    raw = str(value).strip()
    try:
        return int(raw)
    except ValueError:
        try:
            return int(float(raw.replace(" ", "").replace(",", ".")))
        except ValueError as exc:
            raise ValueError(f"invalid integer value: {value}") from exc


def _to_date(value: Any) -> date:
    if value is None:
        raise ValueError("date is missing")
    raw = str(value).strip()
    return date.fromisoformat(raw[:10])


def _discover_files(folder: Path, file_glob: str, filename_regex: str) -> tuple[list[ParsedFile], list[FolderFileReport]]:
    parsed: list[ParsedFile] = []
    reports: list[FolderFileReport] = []
    matcher = re.compile(filename_regex)
    for file_path in sorted(folder.glob(file_glob)):
        if not file_path.is_file():
            continue
        m = matcher.match(file_path.name)
        if not m:
            reports.append(
                FolderFileReport(
                    file_name=file_path.name,
                    status="skipped",
                    error="filename does not match filename_regex",
                )
            )
            continue
        user_email = (m.groupdict().get("user_email") or "").strip().lower()
        dataset = (m.groupdict().get("dataset") or "").strip().lower() or None
        if not user_email:
            reports.append(
                FolderFileReport(
                    file_name=file_path.name,
                    status="skipped",
                    error="regex must provide user_email group",
                )
            )
            continue
        parsed.append(ParsedFile(path=file_path, user_email=user_email, dataset=dataset))
    return parsed, reports


def _pick_value(row: dict[str, Any], aliases: list[str]) -> Any:
    for alias in aliases:
        if alias in row and row.get(alias) not in (None, ""):
            return row.get(alias)
    return None


def _normalize_sales_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "date": _pick_value(row, ["date", "Дата", "dt", "date_from"]),
        "nm_id": _pick_value(row, ["nm_id", "nmId", "Артикул", "nm"]),
        "doc_type": _pick_value(row, ["doc_type", "docType", "Тип документа", "Тип"]),
        "retail_price": _pick_value(row, ["retail_price", "retailPrice", "Розничная цена"]),
        "ppvz_for_pay": _pick_value(row, ["ppvz_for_pay", "ppvzForPay", "К перечислению"]),
        "delivery_rub": _pick_value(row, ["delivery_rub", "Логистика"]),
        "penalty": _pick_value(row, ["penalty", "Штраф"]),
        "additional_payment": _pick_value(row, ["additional_payment", "Доплата"]),
        "storage_fee": _pick_value(row, ["storage_fee", "Хранение"]),
        "quantity": _pick_value(row, ["quantity", "Количество", "Кол-во"]),
    }


def _normalize_ads_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "date": _pick_value(row, ["date", "Дата", "dt", "updTime", "upd_time"]),
        "nm_id": _pick_value(row, ["nm_id", "nmId", "Артикул", "nm"]),
        "campaign_id": _pick_value(row, ["campaign_id", "campaignId", "Кампания", "advertId"]),
        "spend": _pick_value(row, ["spend", "Расход", "updSum"]),
    }


def _read_xlsx_sheet_rows(file_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("openpyxl is required for xlsx import") from exc

    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    headers_row = next(rows_iter, None)
    if not headers_row:
        return []
    headers = [str(h).strip() if h is not None else "" for h in headers_row]
    data: list[dict[str, Any]] = []
    for values in rows_iter:
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}
        if any(v not in (None, "") for v in row.values()):
            data.append(row)
    return data


def _import_sales(db: Session, user_id: str, rows: list[dict[str, str]]) -> int:
    if not rows:
        return 0
    sales: list[RawSale] = []
    min_date: date | None = None
    max_date: date | None = None
    for row in rows:
        try:
            d = _to_date(row.get("date"))
        except ValueError:
            continue
        min_date = d if min_date is None or d < min_date else min_date
        max_date = d if max_date is None or d > max_date else max_date
        sales.append(
            RawSale(
                user_id=user_id,
                date=d,
                nm_id=_to_int(row.get("nm_id"), default=0) or 0,
                doc_type=(row.get("doc_type") or "").strip() or None,
                retail_price=_to_float(row.get("retail_price")),
                ppvz_for_pay=_to_float(row.get("ppvz_for_pay")),
                delivery_rub=_to_float(row.get("delivery_rub")),
                penalty=_to_float(row.get("penalty")),
                additional_payment=_to_float(row.get("additional_payment")),
                storage_fee=_to_float(row.get("storage_fee")),
                quantity=_to_int(row.get("quantity"), default=1) or 1,
            )
        )
    if min_date is None or max_date is None:
        return 0
    db.execute(
        delete(RawSale).where(
            RawSale.user_id == user_id,
            RawSale.date >= min_date,
            RawSale.date <= max_date,
        )
    )
    db.add_all(sales)
    return len(sales)


def _import_ads(db: Session, user_id: str, rows: list[dict[str, str]]) -> int:
    if not rows:
        return 0
    ads: list[RawAd] = []
    min_date: date | None = None
    max_date: date | None = None
    for row in rows:
        try:
            d = _to_date(row.get("date"))
        except ValueError:
            continue
        min_date = d if min_date is None or d < min_date else min_date
        max_date = d if max_date is None or d > max_date else max_date
        ads.append(
            RawAd(
                user_id=user_id,
                date=d,
                nm_id=_to_int(row.get("nm_id"), default=None),
                campaign_id=_to_int(row.get("campaign_id"), default=None),
                spend=_to_float(row.get("spend")),
            )
        )
    if min_date is None or max_date is None:
        return 0
    db.execute(
        delete(RawAd).where(
            RawAd.user_id == user_id,
            RawAd.date >= min_date,
            RawAd.date <= max_date,
        )
    )
    db.add_all(ads)
    return len(ads)


def run_folder_migration(db: Session, current_user: User, body: FolderMigrationRequest) -> FolderMigrationResponse:
    folder = Path(body.folder_path)
    if not folder.exists() or not folder.is_dir():
        raise ValueError("folder_path does not exist or is not a directory")

    discovered, pre_reports = _discover_files(folder, body.file_glob, body.filename_regex)
    reports: list[FolderFileReport] = list(pre_reports)

    total_files = len([p for p in folder.glob(body.file_glob) if p.is_file()])
    matched_files = len(discovered)
    processed_files = 0
    source_rows = 0
    inserted_rows = 0
    rejected_rows = 0
    created_users = 0

    for f in discovered:
        if not body.include_all_users and f.user_email != (current_user.email or "").strip().lower():
            reports.append(
                FolderFileReport(
                    file_name=f.path.name,
                    user_email=f.user_email,
                    dataset=f.dataset,
                    status="skipped",
                    error="file user_email does not match current user",
                )
            )
            continue

        user = db.query(User).filter(User.email == f.user_email).first()
        if not user:
            if not body.auto_create_users:
                reports.append(
                    FolderFileReport(
                        file_name=f.path.name,
                        user_email=f.user_email,
                        dataset=f.dataset,
                        status="rejected",
                        error="user not found by email",
                    )
                )
                continue
            created_users += 1
            if body.dry_run:
                user = User(
                    email=f.user_email,
                    password_hash="dry-run-user",
                    wb_api_key=None,
                    is_active=body.auto_create_users_is_active,
                )
            else:
                raw_password = body.auto_create_users_password or secrets.token_urlsafe(18)
                user = User(
                    email=f.user_email,
                    password_hash=hash_password(raw_password),
                    wb_api_key=None,
                    is_active=body.auto_create_users_is_active,
                )
                db.add(user)
                db.flush()

        try:
            suffix = f.path.suffix.lower()
            if suffix == ".csv":
                with f.path.open("r", encoding=body.encoding, newline="") as fh:
                    reader = csv.DictReader(fh, delimiter=body.delimiter)
                    rows = [dict(r) for r in reader]
                file_source_rows = len(rows)
                file_inserted = 0
                dataset = f.dataset or ""
                if not body.dry_run:
                    if dataset == "sales":
                        file_inserted = _import_sales(db, str(user.id), [_normalize_sales_row(r) for r in rows])
                    elif dataset == "ads":
                        file_inserted = _import_ads(db, str(user.id), [_normalize_ads_row(r) for r in rows])
                    else:
                        raise ValueError("for CSV, dataset group in filename must be sales or ads")
                    db.commit()
                processed_files += 1
                source_rows += file_source_rows
                inserted_rows += file_inserted
                reports.append(
                    FolderFileReport(
                        file_name=f.path.name,
                        user_email=f.user_email,
                        dataset=dataset or None,
                        source_rows=file_source_rows,
                        inserted_rows=file_inserted,
                        status="validated" if body.dry_run else "imported",
                    )
                )
            elif suffix == ".xlsx":
                sales_rows_raw = _read_xlsx_sheet_rows(f.path, "DB_Raw_Data")
                ads_rows_raw = _read_xlsx_sheet_rows(f.path, "DB_Ads_Raw")
                sales_rows = [_normalize_sales_row(r) for r in sales_rows_raw]
                ads_rows = [_normalize_ads_row(r) for r in ads_rows_raw]
                file_source_rows = len(sales_rows) + len(ads_rows)
                sales_inserted = 0
                ads_inserted = 0
                if not body.dry_run:
                    sales_inserted = _import_sales(db, str(user.id), sales_rows)
                    ads_inserted = _import_ads(db, str(user.id), ads_rows)
                    db.commit()
                processed_files += 1
                source_rows += file_source_rows
                inserted_rows += sales_inserted + ads_inserted
                reports.append(
                    FolderFileReport(
                        file_name=f"{f.path.name}#DB_Raw_Data",
                        user_email=f.user_email,
                        dataset="sales",
                        source_rows=len(sales_rows),
                        inserted_rows=sales_inserted,
                        status="validated" if body.dry_run else "imported",
                    )
                )
                reports.append(
                    FolderFileReport(
                        file_name=f"{f.path.name}#DB_Ads_Raw",
                        user_email=f.user_email,
                        dataset="ads",
                        source_rows=len(ads_rows),
                        inserted_rows=ads_inserted,
                        status="validated" if body.dry_run else "imported",
                    )
                )
            else:
                raise ValueError("supported file formats are .csv and .xlsx")
        except (ValueError, OSError, csv.Error) as exc:
            db.rollback()
            rejected_rows += 1
            reports.append(
                FolderFileReport(
                    file_name=f.path.name,
                    user_email=f.user_email,
                    dataset=f.dataset,
                    status="failed",
                    error=str(exc),
                )
            )

    return FolderMigrationResponse(
        dry_run=body.dry_run,
        total_files=total_files,
        matched_files=matched_files,
        processed_files=processed_files,
        source_rows=source_rows,
        inserted_rows=inserted_rows,
        rejected_rows=rejected_rows,
        created_users=created_users,
        files=reports,
    )
