from __future__ import annotations

from collections.abc import Generator
from datetime import date, timedelta
from functools import lru_cache
import uuid
from unittest.mock import MagicMock

import pytest
from fastapi.testclient import TestClient
from sqlalchemy.exc import IntegrityError
from sqlalchemy import text

from app.db import SessionLocal, engine
from app.dependencies import get_current_user
from app.main import app
from app.models.ai_hypothesis import AiHypothesis
from app.models.ai_task import AiTask
from app.models.base import Base
from app.models.sku_daily import SkuDaily
from app.models.user import User


@lru_cache(maxsize=1)
def _ensure_ai_module_schema() -> None:
    """
    These API tests use SessionLocal (real DB). If the local dev DB already has older tables,
    SQLAlchemy create_all() will NOT add missing columns/constraints. We run Alembic upgrade to head
    once to keep schema in sync with the models used in tests.

    Important: local DB may already have ai_* tables created earlier without Alembic stamping.
    We therefore apply minimal additive DDL with IF NOT EXISTS to avoid destructive drops.
    """
    ddl = [
        # fingerprint columns
        "ALTER TABLE ai_tasks ADD COLUMN IF NOT EXISTS fingerprint VARCHAR(80)",
        "ALTER TABLE ai_hypotheses ADD COLUMN IF NOT EXISTS fingerprint VARCHAR(80)",
        # dedupe keys
        "ALTER TABLE ai_tasks ADD COLUMN IF NOT EXISTS dedupe_key VARCHAR(120)",
        "ALTER TABLE ai_hypotheses ADD COLUMN IF NOT EXISTS dedupe_key VARCHAR(120)",
        # indexes
        "CREATE INDEX IF NOT EXISTS ix_ai_tasks_fingerprint ON ai_tasks (fingerprint)",
        "CREATE INDEX IF NOT EXISTS ix_ai_hypotheses_fingerprint ON ai_hypotheses (fingerprint)",
        "CREATE INDEX IF NOT EXISTS ix_ai_tasks_dedupe_key ON ai_tasks (dedupe_key)",
        "CREATE INDEX IF NOT EXISTS ix_ai_hypotheses_dedupe_key ON ai_hypotheses (dedupe_key)",
        # unique per user (use unique index to allow IF NOT EXISTS)
        "CREATE UNIQUE INDEX IF NOT EXISTS ux_ai_tasks_user_fingerprint ON ai_tasks (user_id, fingerprint)",
        "CREATE UNIQUE INDEX IF NOT EXISTS ux_ai_hypotheses_user_fingerprint ON ai_hypotheses (user_id, fingerprint)",
        # open/active unique indexes (partial)
        "CREATE UNIQUE INDEX IF NOT EXISTS ux_ai_tasks_user_dedupe_key_open ON ai_tasks (user_id, dedupe_key) WHERE dedupe_key is not null AND status in ('new','in_progress')",
        "CREATE UNIQUE INDEX IF NOT EXISTS ux_ai_hypotheses_user_dedupe_key_active ON ai_hypotheses (user_id, dedupe_key) WHERE dedupe_key is not null AND status in ('draft','running')",
        # daily log table
        """
        CREATE TABLE IF NOT EXISTS ai_hypothesis_daily_log (
            id UUID NOT NULL,
            hypothesis_id UUID NOT NULL REFERENCES ai_hypotheses(id) ON DELETE CASCADE,
            day DATE NOT NULL,
            happened TEXT NULL,
            changed TEXT NULL,
            unchanged TEXT NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            CONSTRAINT pk_ai_hypothesis_daily_log PRIMARY KEY (id)
        )
        """,
        "CREATE INDEX IF NOT EXISTS ix_ai_hypothesis_daily_log_hypothesis_id ON ai_hypothesis_daily_log (hypothesis_id)",
        "CREATE UNIQUE INDEX IF NOT EXISTS ux_ai_hypothesis_daily_log_hyp_day ON ai_hypothesis_daily_log (hypothesis_id, day)",
        # competitor reports (AI-MVP2)
        """
        CREATE TABLE IF NOT EXISTS ai_competitor_comparison_reports (
            id UUID NOT NULL,
            user_id UUID NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            report_date DATE NOT NULL,
            period VARCHAR(16) NOT NULL DEFAULT 'unknown',
            source VARCHAR(32) NOT NULL DEFAULT 'manual',
            raw_payload JSONB NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            CONSTRAINT pk_ai_competitor_comparison_reports PRIMARY KEY (id)
        )
        """,
        "CREATE INDEX IF NOT EXISTS ix_ai_competitor_comparison_reports_user_id ON ai_competitor_comparison_reports (user_id)",
        "CREATE INDEX IF NOT EXISTS ix_ai_competitor_comparison_reports_report_date ON ai_competitor_comparison_reports (report_date)",
        "CREATE UNIQUE INDEX IF NOT EXISTS ux_ai_competitor_report_user_date_period ON ai_competitor_comparison_reports (user_id, report_date, period)",
        """
        CREATE TABLE IF NOT EXISTS ai_competitor_metrics (
            id UUID NOT NULL,
            report_id UUID NOT NULL REFERENCES ai_competitor_comparison_reports(id) ON DELETE CASCADE,
            nm_id INT NOT NULL,
            metric_code VARCHAR(32) NOT NULL,
            our_value NUMERIC(18,6) NULL,
            competitor_median_value NUMERIC(18,6) NULL,
            unit VARCHAR(16) NULL,
            extra JSONB NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            CONSTRAINT pk_ai_competitor_metrics PRIMARY KEY (id)
        )
        """,
        "CREATE INDEX IF NOT EXISTS ix_ai_competitor_metrics_report_id ON ai_competitor_metrics (report_id)",
        "CREATE INDEX IF NOT EXISTS ix_ai_competitor_metrics_nm_id ON ai_competitor_metrics (nm_id)",
        "CREATE INDEX IF NOT EXISTS ix_ai_competitor_metrics_metric_code ON ai_competitor_metrics (metric_code)",
        # WB cabinet credentials (encrypted)
        """
        CREATE TABLE IF NOT EXISTS ai_wb_cabinet_credentials (
            id UUID NOT NULL,
            user_id UUID NOT NULL UNIQUE REFERENCES users(id) ON DELETE CASCADE,
            wb_login_enc TEXT NOT NULL,
            wb_password_enc TEXT NOT NULL,
            status VARCHAR(16) NOT NULL DEFAULT 'active',
            last_error TEXT NULL,
            last_verified_at TIMESTAMPTZ NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            CONSTRAINT pk_ai_wb_cabinet_credentials PRIMARY KEY (id)
        )
        """,
        "CREATE INDEX IF NOT EXISTS ix_ai_wb_cabinet_credentials_user_id ON ai_wb_cabinet_credentials (user_id)",
        # report lifecycle columns (best-effort additive)
        "ALTER TABLE ai_competitor_comparison_reports ADD COLUMN IF NOT EXISTS valid_until DATE",
        "ALTER TABLE ai_competitor_comparison_reports ADD COLUMN IF NOT EXISTS status VARCHAR(16) NOT NULL DEFAULT 'ready'",
        "ALTER TABLE ai_competitor_comparison_reports ADD COLUMN IF NOT EXISTS cost_or_limit_spent BOOLEAN NOT NULL DEFAULT false",
        "ALTER TABLE ai_competitor_comparison_reports ADD COLUMN IF NOT EXISTS last_error TEXT",
        # actions log
        """
        CREATE TABLE IF NOT EXISTS ai_competitor_report_actions (
            id UUID NOT NULL,
            user_id UUID NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            report_id UUID NULL REFERENCES ai_competitor_comparison_reports(id) ON DELETE SET NULL,
            action VARCHAR(16) NOT NULL,
            result VARCHAR(16) NOT NULL DEFAULT 'ok',
            error_message TEXT NULL,
            requested_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            CONSTRAINT pk_ai_competitor_report_actions PRIMARY KEY (id)
        )
        """,
        "CREATE INDEX IF NOT EXISTS ix_ai_competitor_report_actions_user_id ON ai_competitor_report_actions (user_id)",
        # Competitor metrics: accumulate imports (import_batch_id + latest on report)
        "ALTER TABLE ai_competitor_metrics ADD COLUMN IF NOT EXISTS import_batch_id UUID",
        "ALTER TABLE ai_competitor_comparison_reports ADD COLUMN IF NOT EXISTS latest_import_batch_id UUID",
        """
        DO $$
        DECLARE r RECORD;
        DECLARE bid uuid;
        BEGIN
          FOR r IN SELECT DISTINCT report_id FROM ai_competitor_metrics WHERE import_batch_id IS NULL LOOP
            bid := gen_random_uuid();
            UPDATE ai_competitor_metrics SET import_batch_id = bid
              WHERE report_id = r.report_id AND import_batch_id IS NULL;
            UPDATE ai_competitor_comparison_reports SET latest_import_batch_id = bid WHERE id = r.report_id;
          END LOOP;
        END $$;
        """,
        "DROP INDEX IF EXISTS ux_ai_comp_metric_report_nm_code",
        "ALTER TABLE ai_competitor_metrics DROP CONSTRAINT IF EXISTS uq_ai_comp_metric_report_nm_code",
        "ALTER TABLE ai_competitor_metrics DROP CONSTRAINT IF EXISTS uq_ai_comp_metric_report_nm_code_batch",
        "DROP INDEX IF EXISTS ux_ai_comp_metric_report_nm_code_batch",
        "CREATE UNIQUE INDEX IF NOT EXISTS ux_ai_comp_metric_report_nm_code_batch ON ai_competitor_metrics (report_id, nm_id, metric_code, import_batch_id)",
        "CREATE INDEX IF NOT EXISTS ix_ai_competitor_metrics_import_batch_id ON ai_competitor_metrics (import_batch_id)",
        "ALTER TABLE ai_competitor_metrics ALTER COLUMN import_batch_id SET NOT NULL",
    ]
    with engine.begin() as conn:
        for stmt in ddl:
            conn.execute(text(stmt))


@pytest.fixture
def client() -> Generator[TestClient, None, None]:
    _ensure_ai_module_schema()
    # Ensure tables exist for local runs (CI/dev).
    Base.metadata.create_all(bind=engine)

    user_id = "00000000-0000-0000-0000-000000000111"
    app.dependency_overrides[get_current_user] = lambda: MagicMock(
        id=user_id,
        is_active=True,
        is_admin=True,
    )
    # Store scoping for /ai/*: map store_owner to the same user_id in tests
    from app.dependencies import get_store_context  # local import to avoid import-time cycles

    app.dependency_overrides[get_store_context] = lambda: MagicMock(
        store_owner=MagicMock(id=user_id),
    )

    db = SessionLocal()
    try:
        existing = db.query(User).filter(User.id == user_id).first()
        if not existing:
            db.add(
                User(
                    id=user_id,
                    email="ai-module@example.com",
                    password_hash="x",
                    wb_api_key=None,
                    is_admin=True,
                    is_active=True,
                )
            )
            db.commit()
    finally:
        db.close()

    with TestClient(app) as c:
        yield c

    app.dependency_overrides.pop(get_current_user, None)
    app.dependency_overrides.pop(get_store_context, None)


def _seed_task(user_id: str) -> str:
    db = SessionLocal()
    try:
        t = AiTask(
            user_id=user_id,
            nm_id=123,
            task_type="restock",
            title="Дозакупить товар",
            description="Остатка хватит < 14 дней",
            reason="stock_days_left < 14",
            priority=10,
            status="new",
            fingerprint=f"task:restock:123:{uuid.uuid4()}",
        )
        db.add(t)
        db.commit()
        db.refresh(t)
        return str(t.id)
    finally:
        db.close()


def _seed_hypothesis(user_id: str) -> str:
    db = SessionLocal()
    try:
        h = AiHypothesis(
            user_id=user_id,
            nm_id=123,
            hypothesis_type="content_change",
            title="Поменять контент",
            description="Воронка ниже медианы",
            status="draft",
            fingerprint=f"hyp:content_change:123:{uuid.uuid4()}",
        )
        db.add(h)
        db.commit()
        db.refresh(h)
        return str(h.id)
    finally:
        db.close()


def test_ai_tasks_list_and_patch(client: TestClient) -> None:
    user_id = "00000000-0000-0000-0000-000000000111"
    task_id = _seed_task(user_id)

    r = client.get("/ai/tasks")
    assert r.status_code == 200
    data = r.json()
    assert isinstance(data["items"], list)
    assert any(x["id"] == task_id for x in data["items"])

    r2 = client.patch(f"/ai/tasks/{task_id}", json={"status": "in_progress"})
    assert r2.status_code == 200
    assert r2.json()["status"] == "in_progress"
    assert r2.json()["started_at"] is not None

    r3 = client.patch(f"/ai/tasks/{task_id}", json={"status": "completed"})
    assert r3.status_code == 200
    assert r3.json()["status"] == "completed"
    assert r3.json()["completed_at"] is not None


def test_ai_task_can_complete_from_new(client: TestClient) -> None:
    user_id = "00000000-0000-0000-0000-000000000111"
    task_id = _seed_task(user_id)

    r = client.patch(f"/ai/tasks/{task_id}", json={"status": "completed"})
    assert r.status_code == 200
    assert r.json()["status"] == "completed"
    assert r.json()["started_at"] is not None
    assert r.json()["completed_at"] is not None


def test_wb_access_grant_task_cannot_complete_without_storage_state(client: TestClient) -> None:
    user_id = "00000000-0000-0000-0000-000000000111"
    db = SessionLocal()
    try:
        # Ensure no open wb_access_grant task exists (other tests may have created it).
        db.query(AiTask).filter(
            AiTask.user_id == user_id,
            AiTask.dedupe_key == "task:wb_access_grant",
            AiTask.status.in_(["new", "in_progress"]),
        ).delete()
        db.commit()

        t = AiTask(
            user_id=user_id,
            nm_id=None,
            task_type="wb_access_grant",
            title="Дать доступ к кабинету WB",
            description="test",
            reason=None,
            priority=100,
            status="new",
            fingerprint=None,
            dedupe_key="task:wb_access_grant",
        )
        db.add(t)
        db.commit()
        db.refresh(t)
        task_id = str(t.id)
    finally:
        db.rollback()
        db.close()

    from app.services.ai_wb_access_service import user_storage_state_path

    p = user_storage_state_path(user_id=user_id)
    if p.is_file():
        p.unlink()

    r = client.patch(f"/ai/tasks/{task_id}", json={"status": "completed"})
    assert r.status_code == 409


def test_ai_task_invalid_transition_returns_409(client: TestClient) -> None:
    user_id = "00000000-0000-0000-0000-000000000111"
    task_id = _seed_task(user_id)

    # new -> in_progress is required before cancelled -> completed (invalid)
    r1 = client.patch(f"/ai/tasks/{task_id}", json={"status": "cancelled"})
    assert r1.status_code == 200

    r = client.patch(f"/ai/tasks/{task_id}", json={"status": "completed"})
    assert r.status_code == 409


def test_ai_hypothesis_start_and_finish(client: TestClient) -> None:
    user_id = "00000000-0000-0000-0000-000000000111"
    hypothesis_id = _seed_hypothesis(user_id)

    r0 = client.get("/ai/hypotheses")
    assert r0.status_code == 200
    assert any(x["id"] == hypothesis_id for x in r0.json()["items"])

    r1 = client.post(f"/ai/hypotheses/{hypothesis_id}/start")
    assert r1.status_code == 200

    r2 = client.get(f"/ai/hypotheses/{hypothesis_id}")
    assert r2.status_code == 200
    assert r2.json()["status"] == "running"
    assert r2.json()["started_at"] is not None

    r3 = client.post(f"/ai/hypotheses/{hypothesis_id}/finish", json={"result_summary": "ok"})
    assert r3.status_code == 200

    r4 = client.get(f"/ai/hypotheses/{hypothesis_id}")
    assert r4.status_code == 200
    assert r4.json()["status"] == "finished"
    assert r4.json()["ended_at"] is not None
    assert r4.json()["result_summary"] == "ok"


def test_ai_hypothesis_start_twice_returns_409(client: TestClient) -> None:
    user_id = "00000000-0000-0000-0000-000000000111"
    hypothesis_id = _seed_hypothesis(user_id)

    r1 = client.post(f"/ai/hypotheses/{hypothesis_id}/start")
    assert r1.status_code == 200

    r2 = client.post(f"/ai/hypotheses/{hypothesis_id}/start")
    assert r2.status_code == 409


def test_ai_hypothesis_daily_log_requires_running(client: TestClient) -> None:
    user_id = "00000000-0000-0000-0000-000000000111"
    hypothesis_id = _seed_hypothesis(user_id)

    # draft -> daily log not allowed
    r0 = client.post(
        f"/ai/hypotheses/{hypothesis_id}/daily-log",
        json={"day": "2026-05-10", "happened": "x", "changed": "y", "unchanged": "z"},
    )
    assert r0.status_code == 409

    r1 = client.post(f"/ai/hypotheses/{hypothesis_id}/start")
    assert r1.status_code == 200

    r2 = client.post(
        f"/ai/hypotheses/{hypothesis_id}/daily-log",
        json={"day": "2026-05-10", "happened": "h1", "changed": "c1", "unchanged": "u1"},
    )
    assert r2.status_code == 200
    items = r2.json()["items"]
    assert len(items) == 1
    assert items[0]["day"] == "2026-05-10"
    assert items[0]["happened"] == "h1"

    # Upsert same day -> update, still 1 item
    r3 = client.post(
        f"/ai/hypotheses/{hypothesis_id}/daily-log",
        json={"day": "2026-05-10", "happened": "h2", "changed": "c2", "unchanged": "u2"},
    )
    assert r3.status_code == 200
    items2 = r3.json()["items"]
    assert len(items2) == 1
    assert items2[0]["happened"] == "h2"

    r_get = client.get(f"/ai/hypotheses/{hypothesis_id}/daily-log")
    assert r_get.status_code == 200
    get_items = r_get.json()["items"]
    assert len(get_items) == 1
    assert get_items[0]["day"] == "2026-05-10"
    assert get_items[0]["happened"] == "h2"
    assert get_items[0]["changed"] == "c2"


def test_ai_hypothesis_daily_log_get_empty_for_draft(client: TestClient) -> None:
    user_id = "00000000-0000-0000-0000-000000000111"
    hypothesis_id = _seed_hypothesis(user_id)
    r = client.get(f"/ai/hypotheses/{hypothesis_id}/daily-log")
    assert r.status_code == 200
    assert r.json()["items"] == []


def test_ai_hypothesis_daily_log_get_404_for_unknown(client: TestClient) -> None:
    r = client.get("/ai/hypotheses/ffffffff-ffff-ffff-ffff-ffffffffffff/daily-log")
    assert r.status_code == 404


def test_ai_fingerprint_unique_per_user(client: TestClient) -> None:
    user_id = "00000000-0000-0000-0000-000000000111"
    db = SessionLocal()
    try:
        f = f"dup:{uuid.uuid4()}"
        t1 = AiTask(user_id=user_id, task_type="x", title="t", priority=0, status="new", fingerprint=f)
        t2 = AiTask(user_id=user_id, task_type="y", title="t2", priority=0, status="new", fingerprint=f)
        db.add(t1)
        db.commit()
        db.add(t2)
        with pytest.raises(IntegrityError):
            db.commit()

        db.rollback()
        fh = f"dup_h:{uuid.uuid4()}"
        h1 = AiHypothesis(
            user_id=user_id,
            hypothesis_type="a",
            title="h",
            status="draft",
            fingerprint=fh,
        )
        h2 = AiHypothesis(
            user_id=user_id,
            hypothesis_type="b",
            title="h2",
            status="draft",
            fingerprint=fh,
        )
        db.add(h1)
        db.commit()
        db.add(h2)
        with pytest.raises(IntegrityError):
            db.commit()
    finally:
        db.rollback()
        db.close()


def test_ai_competitor_report_import_list_and_get(client: TestClient) -> None:
    body = {
        "report_date": "2026-05-10",
        "period": "week",
        "source": "manual",
        "raw_payload": {"note": "manual import"},
        "items": [
            {"nm_id": 123, "metric_code": "ctr", "our_value": 3.1, "competitor_median_value": 4.2, "unit": "%"},
            {"nm_id": 123, "metric_code": "traffic", "our_value": 1000, "competitor_median_value": 1500},
        ],
    }
    r = client.post("/ai/competitor-reports/import", json=body)
    assert r.status_code == 200
    rep_id = r.json()["id"]
    assert r.json()["period"] == "week"

    r2 = client.get("/ai/competitor-reports")
    assert r2.status_code == 200
    assert any(x["id"] == rep_id for x in r2.json()["items"])

    r3 = client.get(f"/ai/competitor-reports/{rep_id}")
    assert r3.status_code == 200
    data = r3.json()
    assert data["report"]["id"] == rep_id
    assert data.get("raw_payload") == {"note": "manual import"}
    metrics = data["metrics"]
    assert len(metrics) == 2
    assert any(m["metric_code"] == "ctr" for m in metrics)


def test_ai_competitor_report_import_is_idempotent_by_date_and_period(client: TestClient) -> None:
    base = {
        "report_date": "2026-05-10",
        "period": "week",
        "source": "manual",
        "items": [{"nm_id": 123, "metric_code": "ctr", "our_value": 1.0, "competitor_median_value": 2.0}],
    }
    r1 = client.post("/ai/competitor-reports/import", json=base)
    assert r1.status_code == 200
    rid1 = r1.json()["id"]

    base2 = {
        **base,
        "items": [{"nm_id": 123, "metric_code": "ctr", "our_value": 5.0, "competitor_median_value": 6.0}],
    }
    r2 = client.post("/ai/competitor-reports/import", json=base2)
    assert r2.status_code == 200
    rid2 = r2.json()["id"]
    assert rid2 == rid1

    r3 = client.get(f"/ai/competitor-reports/{rid1}")
    assert r3.status_code == 200
    metrics = r3.json()["metrics"]
    assert len(metrics) == 1
    assert float(metrics[0]["our_value"]) == 5.0


def test_ai_competitor_report_reimport_accumulates_metric_batches(client: TestClient) -> None:
    from app.db import SessionLocal
    from app.models.ai_competitor_report import AiCompetitorComparisonReport

    user_id = "00000000-0000-0000-0000-000000000111"
    iso_day = "2035-07-07"
    nm_isolated = 66077331
    db = SessionLocal()
    try:
        db.query(AiCompetitorComparisonReport).filter(
            AiCompetitorComparisonReport.user_id == user_id,
            AiCompetitorComparisonReport.report_date == date.fromisoformat(iso_day),
            AiCompetitorComparisonReport.period == "week",
        ).delete()
        db.commit()
    finally:
        db.close()

    base = {
        "report_date": iso_day,
        "period": "week",
        "source": "manual",
        "items": [{"nm_id": nm_isolated, "metric_code": "ctr", "our_value": 1.0, "competitor_median_value": 2.0}],
    }
    r1 = client.post("/ai/competitor-reports/import", json=base)
    assert r1.status_code == 200
    rid = r1.json()["id"]

    r2 = client.post(
        "/ai/competitor-reports/import",
        json={
            **base,
            "items": [{"nm_id": nm_isolated, "metric_code": "ctr", "our_value": 9.0, "competitor_median_value": 10.0}],
        },
    )
    assert r2.status_code == 200
    assert r2.json()["id"] == rid

    r_latest = client.get(f"/ai/competitor-reports/{rid}")
    assert r_latest.status_code == 200
    assert len(r_latest.json()["metrics"]) == 1
    assert float(r_latest.json()["metrics"][0]["our_value"]) == 9.0

    r_all = client.get(f"/ai/competitor-reports/{rid}", params={"metrics_scope": "all"})
    assert r_all.status_code == 200
    ctr_rows = [m for m in r_all.json()["metrics"] if m["metric_code"] == "ctr" and m["nm_id"] == nm_isolated]
    assert len(ctr_rows) == 2
    bids = {m["import_batch_id"] for m in ctr_rows}
    assert len(bids) == 2


def test_ai_competitor_report_import_rejects_invalid_metric_code(client: TestClient) -> None:
    body = {
        "report_date": "2026-05-10",
        "period": "week",
        "source": "manual",
        "items": [{"nm_id": 123, "metric_code": "bad", "our_value": 1}],
    }
    r = client.post("/ai/competitor-reports/import", json=body)
    assert r.status_code == 400


def _seed_sku_daily_logistics_spike(user_id: str, nm_id: int) -> None:
    """
    Creates 2 days of sku_daily with prev=100 and today=140 -> +40% spike (day-to-day).
    """
    db = SessionLocal()
    try:
        base_day = date.fromisoformat("2026-05-10")
        prev_day = base_day - timedelta(days=1)
        db.query(SkuDaily).filter(
            SkuDaily.user_id == user_id,
            SkuDaily.nm_id == nm_id,
            SkuDaily.date >= prev_day,
            SkuDaily.date <= base_day,
        ).delete()
        db.commit()
        db.add(
            SkuDaily(
                user_id=user_id,
                date=prev_day,
                nm_id=nm_id,
                logistics=100,
                revenue=1000,
                margin=200,
                ads_spend=50,
                open_count=100,
                order_count=10,
            )
        )
        db.add(
            SkuDaily(
                user_id=user_id,
                date=base_day,
                nm_id=nm_id,
                logistics=140,
                revenue=1000,
                margin=200,
                ads_spend=50,
                open_count=100,
                order_count=10,
            )
        )
        db.commit()
    finally:
        db.rollback()
        db.close()


def _seed_sku_daily_logistics_spike_avg7d(user_id: str, nm_id: int) -> None:
    """
    Creates 8 days: previous 7 days logistics=100, today=140 -> +40% vs avg7d.
    """
    db = SessionLocal()
    try:
        base_day = date.fromisoformat("2026-05-10")
        start_day = base_day - timedelta(days=7)
        db.query(SkuDaily).filter(
            SkuDaily.user_id == user_id,
            SkuDaily.nm_id == nm_id,
            SkuDaily.date >= start_day,
            SkuDaily.date <= base_day,
        ).delete()
        db.commit()

        for i in range(7):
            d = base_day - timedelta(days=i + 1)
            db.add(
                SkuDaily(
                    user_id=user_id,
                    date=d,
                    nm_id=nm_id,
                    logistics=100,
                    revenue=1000,
                    margin=200,
                    ads_spend=50,
                    open_count=100,
                    order_count=10,
                )
            )
        db.add(
            SkuDaily(
                user_id=user_id,
                date=base_day,
                nm_id=nm_id,
                logistics=140,
                revenue=1000,
                margin=200,
                ads_spend=50,
                open_count=100,
                order_count=10,
            )
        )
        db.commit()
    finally:
        db.rollback()
        db.close()


def test_ai_daily_analytics_run_creates_entities_and_is_idempotent(client: TestClient) -> None:
    user_id = "00000000-0000-0000-0000-000000000111"

    report_date = "2026-05-10"
    period = "week"

    # Clean up first so open dedupe_keys do not block new hypothesis inserts (import accumulates metric batches).
    db = SessionLocal()
    try:
        for fp in (
            f"hyp:content_change:123:{report_date}:{period}",
            f"hyp:ab_test:123:{report_date}:{period}",
            f"task:self_buyouts:123:{report_date}:{period}",
            "task:restock:123:2026-05-10",
            f"task:check_measurements:123:{report_date}:2026-05-10",
            f"task:check_ktr:123:{report_date}:2026-05-10",
        ):
            db.query(AiHypothesis).filter(AiHypothesis.user_id == user_id, AiHypothesis.fingerprint == fp).delete(
                synchronize_session=False
            )
            db.query(AiTask).filter(AiTask.user_id == user_id, AiTask.fingerprint == fp).delete(synchronize_session=False)
        for dk in ("hyp:ab_test:123", "hyp:content_change:123"):
            db.query(AiHypothesis).filter(AiHypothesis.user_id == user_id, AiHypothesis.dedupe_key == dk).delete(
                synchronize_session=False
            )
        db.commit()
    finally:
        db.close()

    # Import competitor report for nm_id=123 with bad funnels + bad ctr
    body = {
        "report_date": report_date,
        "period": period,
        "source": "manual",
        "items": [
            {"nm_id": 123, "metric_code": "ctr", "our_value": 3.1, "competitor_median_value": 4.2, "unit": "%"},
            {"nm_id": 123, "metric_code": "funnel_cart", "our_value": 1.0, "competitor_median_value": 2.0},
        ],
    }
    r = client.post("/ai/competitor-reports/import", json=body)
    assert r.status_code == 200
    rep_id = r.json()["id"]

    _seed_sku_daily_logistics_spike(user_id, 123)

    # Run analytics
    r2 = client.post(
        "/ai/analytics/run",
        json={
            "report_id": rep_id,
            "date_for": "2026-05-10",
            "stock_days_left": {"123": 10},
            "social": {"123": {"reviews": 10, "rating": 4.2}},
        },
    )
    assert r2.status_code == 200
    data = r2.json()
    assert data["status"] == "ok"
    assert data["report_id"] == rep_id
    assert len(data["created_hypothesis_ids"]) >= 2  # content_change + ab_test
    assert len(data["created_task_ids"]) >= 3  # restock + 2 logistics tasks + self_buyouts

    db2 = SessionLocal()
    try:
        fps = (
            f"hyp:content_change:123:{report_date}:{period}",
            f"hyp:ab_test:123:{report_date}:{period}",
        )
        hyps = (
            db2.query(AiHypothesis)
            .filter(AiHypothesis.user_id == user_id, AiHypothesis.fingerprint.in_(fps))
            .all()
        )
        assert len(hyps) == 2
        for h in hyps:
            tr = (h.trigger_reason or "").lower()
            assert "funnel_cart" not in tr
            assert "funnel_order" not in tr
            assert "metric_code" not in tr
            assert "рекомендуется" in tr
    finally:
        db2.close()

    # Second run should create nothing new (fingerprints)
    r3 = client.post("/ai/analytics/run", json={"report_id": rep_id, "date_for": "2026-05-10"})
    assert r3.status_code == 200
    data2 = r3.json()
    assert data2["created_task_ids"] == []
    assert data2["created_hypothesis_ids"] == []


def test_ai_daily_analytics_logistics_rule_uses_avg7d_when_available(client: TestClient) -> None:
    user_id = "00000000-0000-0000-0000-000000000111"

    report_date = "2026-05-10"
    period = "week"
    body = {
        "report_date": report_date,
        "period": period,
        "source": "manual",
        "items": [
            {"nm_id": 123, "metric_code": "ctr", "our_value": 3.1, "competitor_median_value": 4.2, "unit": "%"},
        ],
    }
    r = client.post("/ai/competitor-reports/import", json=body)
    assert r.status_code == 200
    rep_id = r.json()["id"]

    # Cleanup deterministic fingerprints for logistics tasks (so test is repeatable)
    db = SessionLocal()
    try:
        for task_type in ("check_measurements", "check_ktr"):
            fp = f"task:{task_type}:123:{report_date}:2026-05-10"
            db.query(AiTask).filter(AiTask.user_id == user_id, AiTask.fingerprint == fp).delete()
        db.commit()
    finally:
        db.rollback()
        db.close()

    _seed_sku_daily_logistics_spike_avg7d(user_id, 123)

    r2 = client.post("/ai/analytics/run", json={"report_id": rep_id, "date_for": "2026-05-10"})
    assert r2.status_code == 200
    created = r2.json()["created_task_ids"]
    assert len(created) >= 2

def test_ai_daily_analytics_task_dedupe_updates_open_and_creates_after_close(client: TestClient) -> None:
    """
    Rule:
    - if task is open (new|in_progress) -> analytics updates it (no new id)
    - if task is closed (completed|cancelled) -> analytics creates a new row
    """
    user_id = "00000000-0000-0000-0000-000000000111"

    report_date = "2026-05-10"
    period = "week"
    body = {
        "report_date": report_date,
        "period": period,
        "source": "manual",
        "items": [
            {"nm_id": 123, "metric_code": "ctr", "our_value": 3.1, "competitor_median_value": 4.2, "unit": "%"},
        ],
    }
    r = client.post("/ai/competitor-reports/import", json=body)
    assert r.status_code == 200
    rep_id = r.json()["id"]

    # Cleanup deterministic fingerprints
    fp = "task:restock:123:2026-05-10"
    db = SessionLocal()
    try:
        db.query(AiTask).filter(AiTask.user_id == user_id, AiTask.fingerprint == fp).delete()
        # Also cleanup any OPEN task by dedupe_key (test isolation; fingerprints can collide with closed rows)
        db.query(AiTask).filter(
            AiTask.user_id == user_id,
            AiTask.dedupe_key == "task:restock:123",
            AiTask.status.in_(["new", "in_progress"]),
        ).delete()
        db.commit()
    finally:
        db.rollback()
        db.close()

    # 1) First run creates restock task (open)
    r1 = client.post("/ai/analytics/run", json={"report_id": rep_id, "date_for": "2026-05-10", "stock_days_left": {"123": 10}})
    assert r1.status_code == 200
    created1 = r1.json()["created_task_ids"]
    assert len(created1) >= 1

    # 2) Second run updates open task (no new)
    r2 = client.post("/ai/analytics/run", json={"report_id": rep_id, "date_for": "2026-05-10", "stock_days_left": {"123": 9}})
    assert r2.status_code == 200
    assert r2.json()["created_task_ids"] == []

    # Close the open task
    db = SessionLocal()
    try:
        open_row = (
            db.query(AiTask)
            .filter(AiTask.user_id == user_id, AiTask.dedupe_key == "task:restock:123", AiTask.status.in_(["new", "in_progress"]))
            .order_by(AiTask.created_at.desc())
            .first()
        )
        assert open_row is not None
        open_row.status = "completed"
        db.add(open_row)
        db.commit()
    finally:
        db.rollback()
        db.close()

    # 3) Third run should create a new restock task because previous is closed
    r3 = client.post("/ai/analytics/run", json={"report_id": rep_id, "date_for": "2026-05-10", "stock_days_left": {"123": 8}})
    assert r3.status_code == 200
    created3 = r3.json()["created_task_ids"]
    assert len(created3) >= 1


def test_ai_daily_analytics_hypothesis_dedupe_blocks_duplicate_when_active(client: TestClient) -> None:
    """
    Rule: repeated hypotheses are not generated if there is an existing one in status != cancelled/finished
    (i.e. active draft/running exists for the same nm_id + hypothesis_type).
    """
    user_id = "00000000-0000-0000-0000-000000000111"

    report_date = "2026-05-10"
    period = "week"
    body = {
        "report_date": report_date,
        "period": period,
        "source": "manual",
        "items": [
            {"nm_id": 123, "metric_code": "funnel_cart", "our_value": 1.0, "competitor_median_value": 2.0},
        ],
    }
    r = client.post("/ai/competitor-reports/import", json=body)
    assert r.status_code == 200
    rep_id = r.json()["id"]

    # cleanup deterministic fingerprint
    fp = f"hyp:content_change:123:{report_date}:{period}"
    db = SessionLocal()
    try:
        db.query(AiHypothesis).filter(AiHypothesis.user_id == user_id, AiHypothesis.fingerprint == fp).delete()
        db.commit()
    finally:
        db.rollback()
        db.close()

    # first run creates content_change hypothesis (draft)
    r1 = client.post("/ai/analytics/run", json={"report_id": rep_id, "date_for": "2026-05-10"})
    assert r1.status_code == 200
    assert len(r1.json()["created_hypothesis_ids"]) >= 1

    # second run should not create a new one because an active draft exists
    r2 = client.post("/ai/analytics/run", json={"report_id": rep_id, "date_for": "2026-05-10"})
    assert r2.status_code == 200
    assert r2.json()["created_hypothesis_ids"] == []


def test_ai_daily_analytics_run_unknown_report_returns_404(client: TestClient) -> None:
    r = client.post("/ai/analytics/run", json={"report_id": "00000000-0000-0000-0000-000000000999"})
    assert r.status_code == 404


def test_ai_wb_credentials_upsert_and_status(client: TestClient, monkeypatch) -> None:
    from cryptography.fernet import Fernet

    monkeypatch.setenv("APP_ENCRYPTION_KEY", Fernet.generate_key().decode("utf-8"))

    r0 = client.get("/ai/wb-credentials/status")
    assert r0.status_code == 200
    assert r0.json()["status"] in {"missing", "active", "invalid", "needs_reauth", "disabled"}

    r1 = client.put("/ai/wb-credentials", json={"wb_login": "user", "wb_password": "pass"})
    assert r1.status_code == 200
    assert r1.json()["status"] == "active"

    r2 = client.get("/ai/wb-credentials/status")
    assert r2.status_code == 200
    assert r2.json()["status"] == "active"


def test_ai_wb_access_remote_status_proxies_internal_manager(client: TestClient, monkeypatch) -> None:
    class _Resp:
        def __init__(self, body: str) -> None:
            self._body = body.encode("utf-8")

        def read(self) -> bytes:
            return self._body

        def __enter__(self) -> "_Resp":
            return self

        def __exit__(self, exc_type, exc, tb) -> None:  # noqa: ANN001
            return None

    def _fake_urlopen(req, timeout=0):  # noqa: ANN001
        assert "http://wb_auth:8081/status" in str(getattr(req, "full_url", ""))
        return _Resp('{"status":"ok","active":true,"url":"https://seller.wildberries.ru/"}')

    monkeypatch.setenv("WB_AUTH_INTERNAL_TOKEN", "t")
    monkeypatch.setattr("urllib.request.urlopen", _fake_urlopen)

    r = client.post("/ai/wb-access/remote/status")
    assert r.status_code == 200
    data = r.json()
    assert data["status"] == "ok"
    assert data["active"] is True


def test_ai_tasks_list_auto_creates_wb_access_task_when_missing(client: TestClient) -> None:
    """
    UX contract: when WB access (storage_state) is missing, AI module should expose a single
    human-readable task prompting the user to grant access.
    """
    user_id = "00000000-0000-0000-0000-000000000111"
    from app.services.ai_wb_access_service import user_storage_state_path

    p = user_storage_state_path(user_id=user_id)
    if p.is_file():
        p.unlink()

    r = client.get("/ai/tasks")
    assert r.status_code == 200
    items = r.json()["items"]
    assert any(x.get("task_type") == "wb_access_grant" and x.get("status") in {"new", "in_progress"} for x in items)


def test_ai_competitor_report_refresh_flow_creates_task_and_executes(client: TestClient, monkeypatch) -> None:
    from cryptography.fernet import Fernet

    monkeypatch.setenv("APP_ENCRYPTION_KEY", Fernet.generate_key().decode("utf-8"))
    # Disable real Playwright in tests; patch task function to avoid network.
    monkeypatch.setenv("AI_COMPETITOR_PLAYWRIGHT_ENABLED", "0")

    # Test isolation: remove any existing open refresh task
    db = SessionLocal()
    try:
        db.query(AiTask).filter(
            AiTask.user_id == "00000000-0000-0000-0000-000000000111",
            AiTask.dedupe_key == "task:competitor_report_refresh:week",
            AiTask.status.in_(["new", "in_progress"]),
        ).delete()
        db.commit()
    finally:
        db.rollback()
        db.close()

    # Create refresh task (explicit confirmation)
    r1 = client.post("/ai/competitor-reports/request-refresh", json={"period": "week"})
    assert r1.status_code == 200
    task_id = r1.json()["id"]
    assert r1.json()["task_type"] == "competitor_report_refresh"
    assert r1.json()["status"] == "new"

    # Execute should enqueue (but with Playwright disabled it still queues; actual worker handles error)
    r2 = client.post(f"/ai/tasks/{task_id}/execute")
    assert r2.status_code == 200
    assert r2.json()["status"] == "ok"


def test_ai_competitor_report_status_path_not_shadowed_by_report_id_route(client: TestClient) -> None:
    """Регрессия: /competitor-reports/status не должен матчиться как {report_id}=status."""
    r = client.get("/ai/competitor-reports/status", params={"period": "week"})
    assert r.status_code == 200
    data = r.json()
    assert "status" in data
    assert data["status"] in {"missing", "ready", "stale", "running", "error"}


def test_ai_competitor_report_status_falls_back_to_any_period_if_requested_missing(client: TestClient) -> None:
    """
    Regression/UX: UI requests period=week, but user may already have a report saved under another period.
    In that case we should NOT show "missing".
    """
    body = {
        "report_date": "2026-05-10",
        "period": "month",
        "source": "manual",
        "items": [
            {"nm_id": 123, "metric_code": "ctr", "our_value": 3.1, "competitor_median_value": 4.2, "unit": "%"},
        ],
    }
    r0 = client.post("/ai/competitor-reports/import", json=body)
    assert r0.status_code == 200

    r = client.get("/ai/competitor-reports/status", params={"period": "week"})
    assert r.status_code == 200
    data = r.json()
    assert data["status"] in {"ready", "stale", "running", "error"}
    assert data.get("report_id")


def test_ai_competitor_report_actions_list_returns_store_owner_rows(client: TestClient) -> None:
    from app.models.ai_competitor_report_action import AiCompetitorReportAction
    from app.models.base import uuid_gen

    user_id = "00000000-0000-0000-0000-000000000111"
    db = SessionLocal()
    try:
        db.query(AiCompetitorReportAction).filter(AiCompetitorReportAction.user_id == user_id).delete()
        db.commit()
        aid = str(uuid_gen())
        db.add(
            AiCompetitorReportAction(
                id=aid,
                user_id=user_id,
                report_id=None,
                action="refresh",
                result="error",
                error_message="stub failure",
            )
        )
        db.commit()
    finally:
        db.close()

    r = client.get("/ai/competitor-reports/actions")
    assert r.status_code == 200
    items = r.json()["items"]
    assert len(items) >= 1
    assert items[0]["id"] == aid
    assert items[0]["action"] == "refresh"
    assert items[0]["result"] == "error"
    assert items[0]["error_message"] == "stub failure"


def test_competitor_excel_parser_header_not_first_row() -> None:
    """
    Regression: WB excel may have title rows before header.
    Parser must detect header row and start data after it.
    """
    from datetime import date as date_type
    from io import BytesIO

    from openpyxl import Workbook

    from app.services.ai_competitor_excel_parser import parse_wb_competitor_excel

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["Some title"])
    ws.append(["Another line"])
    ws.append([])
    ws.append(["Generated:", "2026-05-10"])
    ws.append(["nm_id", "ctr", "ctr_median"])
    ws.append([123, 3.1, 4.2])

    buf = BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    payload = parse_wb_competitor_excel(
        content=content,
        report_date=date_type.fromisoformat("2026-05-10"),
        period="week",
        raw_payload={"x": 1},
    )
    assert payload["period"] == "week"
    assert payload["source"] == "playwright"
    assert any(x["nm_id"] == 123 and x["metric_code"] == "ctr" for x in payload["items"])


def test_ai_competitor_report_worker_success_sets_ready_and_logs_action(client: TestClient, monkeypatch) -> None:
    from datetime import date as date_type
    from io import BytesIO

    from openpyxl import Workbook
    from sqlalchemy.orm import Session

    from app.db import SessionLocal
    from app.models.ai_competitor_report import AiCompetitorComparisonReport
    from app.models.ai_competitor_report_action import AiCompetitorReportAction
    from app.services.ai_wb_credentials_service import upsert_credentials

    # crypto key for credentials
    from cryptography.fernet import Fernet

    monkeypatch.setenv("APP_ENCRYPTION_KEY", Fernet.generate_key().decode("utf-8"))
    monkeypatch.setenv("AI_COMPETITOR_PLAYWRIGHT_ENABLED", "1")
    monkeypatch.setenv("AI_COMPETITOR_PLAYWRIGHT_FETCH_PERIODS", "week,month")

    # Seed credentials
    user_id = "00000000-0000-0000-0000-000000000111"
    db: Session = SessionLocal()
    try:
        upsert_credentials(db=db, user_id=user_id, wb_login="u", wb_password="p")
        # Clear open dedupe keys for stub nm_id=123 (second period reuses same dedupe_key -> no new ids).
        for dk in ("hyp:ab_test:123", "hyp:content_change:123"):
            db.query(AiHypothesis).filter(AiHypothesis.user_id == user_id, AiHypothesis.dedupe_key == dk).delete(
                synchronize_session=False
            )
        db.commit()
    finally:
        db.close()

    # Build minimal workbook bytes for parser
    wb = Workbook()
    ws = wb.active
    ws.append(["nm_id", "ctr", "ctr_median"])
    ws.append([123, 3.1, 4.2])
    buf = BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    # Wrap into zip to mimic WB export (often provides .zip with .xlsx inside)
    from zipfile import ZipFile
    zbuf = BytesIO()
    with ZipFile(zbuf, "w") as zf:
        zf.writestr("report.xlsx", excel_bytes)
    zip_bytes = zbuf.getvalue()

    # Monkeypatch Playwright fetch to avoid network and selectors
    import app.services.ai_competitor_playwright as pw

    def _fake_fetch_comparison_excel_bytes(*, login: str, password: str, period: str, storage_state_path: str | None = None):
        assert login == "u"
        assert password == "p"
        assert period in {"week", "month"}
        return zip_bytes, {"stub": True, "period": period, "suggested_filename": "wb_export.zip"}

    monkeypatch.setattr(pw, "fetch_comparison_excel_bytes", _fake_fetch_comparison_excel_bytes)

    from celery_app.tasks import ai_competitor_report_fetch_playwright

    res = ai_competitor_report_fetch_playwright(user_id, "week")
    assert res["ok"] is True
    assert isinstance(res.get("reports"), list)
    assert {x.get("period") for x in res["reports"]} >= {"week", "month"}
    for rpt in res["reports"]:
        an = rpt.get("analytics") or {}
        assert an.get("ok") is True
    # week creates new hypothesis id; month hits same dedupe_key -> updates row, often empty created_*.
    total_new_hyps = sum(len((x.get("analytics") or {}).get("created_hypothesis_ids") or []) for x in res["reports"])
    assert total_new_hyps >= 1

    db2: Session = SessionLocal()
    try:
        today = date_type.today()
        for p in ("week", "month"):
            rep = (
                db2.query(AiCompetitorComparisonReport)
                .filter(
                    AiCompetitorComparisonReport.user_id == user_id,
                    AiCompetitorComparisonReport.report_date == today,
                    AiCompetitorComparisonReport.period == p,
                )
                .first()
            )
            assert rep is not None
            assert rep.status == "ready"
            assert rep.last_error is None

            act = (
                db2.query(AiCompetitorReportAction)
                .filter(AiCompetitorReportAction.user_id == user_id, AiCompetitorReportAction.report_id == str(rep.id))
                .order_by(AiCompetitorReportAction.requested_at.desc())
                .first()
            )
            assert act is not None
            assert act.action == "refresh"
            assert act.result == "ok"
    finally:
        db2.close()


def test_ai_competitor_report_worker_can_run_with_storage_state_without_creds(client: TestClient, monkeypatch, tmp_path) -> None:
    """
    If WB_PLAYWRIGHT_STORAGE_STATE_PATH is provided (phone+code auth done manually once),
    worker must be able to run even when encrypted creds are missing/invalid.
    """
    from datetime import date as date_type
    from io import BytesIO

    from openpyxl import Workbook
    from sqlalchemy.orm import Session

    from app.db import SessionLocal
    from app.models.ai_competitor_report import AiCompetitorComparisonReport

    monkeypatch.setenv("AI_COMPETITOR_PLAYWRIGHT_ENABLED", "1")
    monkeypatch.setenv("AI_COMPETITOR_PLAYWRIGHT_FETCH_PERIODS", "week")
    state = tmp_path / "state.json"
    state.write_text('{"cookies":[],"origins":[]}', encoding="utf-8")
    monkeypatch.setenv("WB_PLAYWRIGHT_STORAGE_STATE_PATH", str(state))

    # Build minimal workbook bytes for parser
    wb = Workbook()
    ws = wb.active
    ws.append(["nm_id", "ctr", "ctr_median"])
    ws.append([123, 3.1, 4.2])
    buf = BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    import app.services.ai_competitor_playwright as pw

    def _fake_fetch_comparison_excel_bytes(*, login: str, password: str, period: str, storage_state_path: str | None = None):
        assert login == ""
        assert password == ""
        assert period == "week"
        return excel_bytes, {"stub": True, "period": period}

    monkeypatch.setattr(pw, "fetch_comparison_excel_bytes", _fake_fetch_comparison_excel_bytes)

    from celery_app.tasks import ai_competitor_report_fetch_playwright

    user_id = "00000000-0000-0000-0000-000000000222"
    res = ai_competitor_report_fetch_playwright(user_id, "week")
    assert res["ok"] is True
    assert res.get("reports") and res["reports"][0].get("analytics", {}).get("ok") is True

    db2: Session = SessionLocal()
    try:
        today = date_type.today()
        rep = (
            db2.query(AiCompetitorComparisonReport)
            .filter(
                AiCompetitorComparisonReport.user_id == user_id,
                AiCompetitorComparisonReport.report_date == today,
                AiCompetitorComparisonReport.period == "week",
            )
            .first()
        )
        assert rep is not None
        assert rep.status == "ready"
    finally:
        db2.close()


def test_ai_daily_analytics_does_not_trigger_content_change_on_funnel_median_ceiling_100(
    client: TestClient,
) -> None:
    """Медиана конкурентов ровно 100 п.п. по конверсии — подозрительный кэп; правило content_change не срабатывает."""
    user_id = "00000000-0000-0000-0000-000000000111"
    report_date = "2026-05-12"
    period = "week"
    nm_id = 55501
    fp = f"hyp:content_change:{nm_id}:{report_date}:{period}"

    db = SessionLocal()
    try:
        db.query(AiHypothesis).filter(AiHypothesis.user_id == user_id, AiHypothesis.fingerprint == fp).delete(
            synchronize_session=False
        )
        db.commit()
    finally:
        db.close()

    body = {
        "report_date": report_date,
        "period": period,
        "source": "manual",
        "items": [
            {
                "nm_id": nm_id,
                "metric_code": "funnel_order",
                "our_value": 15.0,
                "competitor_median_value": 100.0,
            },
        ],
    }
    r = client.post("/ai/competitor-reports/import", json=body)
    assert r.status_code == 200
    rep_id = r.json()["id"]

    r2 = client.post("/ai/analytics/run", json={"report_id": rep_id, "date_for": report_date})
    assert r2.status_code == 200
    data = r2.json()
    assert data.get("created_hypothesis_ids") == []

    db2 = SessionLocal()
    try:
        assert db2.query(AiHypothesis).filter(AiHypothesis.user_id == user_id, AiHypothesis.fingerprint == fp).first() is None
    finally:
        db2.close()


def test_ai_self_buyouts_from_imported_review_rows_without_social(client: TestClient) -> None:
    """Строки отчёта «Количество отзывов» / «Рейтинг» в импорте — достаточно для self_buyouts при слабой воронке."""
    user_id = "00000000-0000-0000-0000-000000000111"
    report_date = "2026-05-15"
    period = "week"
    nm_id = 90123
    fp_self = f"task:self_buyouts:{nm_id}:{report_date}:{period}"
    fp_hyp_c = f"hyp:content_change:{nm_id}:{report_date}:{period}"
    fp_hyp_a = f"hyp:ab_test:{nm_id}:{report_date}:{period}"

    db = SessionLocal()
    try:
        db.query(AiTask).filter(AiTask.user_id == user_id, AiTask.fingerprint == fp_self).delete(synchronize_session=False)
        for fp in (fp_hyp_c, fp_hyp_a):
            db.query(AiHypothesis).filter(AiHypothesis.user_id == user_id, AiHypothesis.fingerprint == fp).delete(
                synchronize_session=False
            )
        db.commit()
    finally:
        db.close()

    body = {
        "report_date": report_date,
        "period": period,
        "source": "manual",
        "items": [
            {"nm_id": nm_id, "metric_code": "ctr", "our_value": 3.1, "competitor_median_value": 4.2, "unit": "%"},
            {"nm_id": nm_id, "metric_code": "funnel_cart", "our_value": 1.0, "competitor_median_value": 2.0},
            {"nm_id": nm_id, "metric_code": "review_count", "our_value": 10.0, "competitor_median_value": 200.0},
            {"nm_id": nm_id, "metric_code": "review_rating", "our_value": 4.5, "competitor_median_value": 4.8},
        ],
    }
    r = client.post("/ai/competitor-reports/import", json=body)
    assert r.status_code == 200
    rep_id = r.json()["id"]

    r2 = client.post("/ai/analytics/run", json={"report_id": rep_id, "date_for": report_date})
    assert r2.status_code == 200

    db2 = SessionLocal()
    try:
        row = db2.query(AiTask).filter(AiTask.user_id == user_id, AiTask.fingerprint == fp_self).first()
        assert row is not None
        assert row.task_type == "self_buyouts"
    finally:
        db2.close()

