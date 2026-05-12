from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.services.ai_competitor_excel_parser import parse_wb_competitor_excel


def test_parse_cards_comparison_explicit_conversion_rows_not_scaled_by_100() -> None:
    """WB: проценты в ячейках без «%» — не умножаем на 100 (8 = 8%, не 800%)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Показатели"
    ws.append(["Отчёт"])
    ws.append(["Показатели", "Артикул WB 10", "Артикул WB 20"])
    ws.append(["Показы", 1000, 2000])
    ws.append(["CTR", 3.0, 4.0])
    ws.append(["Конверсия в корзину, %", 8.0, 12.0])
    ws.append(["Конверсия в заказ, %", 2.5, 3.0])
    buf = BytesIO()
    wb.save(buf)

    out = parse_wb_competitor_excel(content=buf.getvalue(), report_date=date(2026, 5, 11), period="week")
    cart_10 = next(i for i in out["items"] if i["nm_id"] == 10 and i["metric_code"] == "funnel_cart")
    assert cart_10["our_value"] == 8.0
    assert cart_10["competitor_median_value"] == 12.0
    ord_20 = next(i for i in out["items"] if i["nm_id"] == 20 and i["metric_code"] == "funnel_order")
    assert ord_20["our_value"] == 3.0
    assert ord_20["competitor_median_value"] == 2.5


def test_parse_cards_comparison_pokazateli_sheet() -> None:
    """
    New WB export: sheet 'Показатели' with nm_id columns and metric rows.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Показатели"
    ws.append(["Отчёт «Сравнение карточек»: Показатели"])
    ws.append(["Показатели", "Артикул WB 111", "Артикул WB 222", "Разница артикул 222 - артикул 111"])
    ws.append(["Показы", 100, 300, 200])
    ws.append(["CTR", 5, 15, 10])
    ws.append(["Добавления в корзину, шт", 10, 20, 10])
    ws.append(["Заказы, шт", 2, 8, 6])
    ws.append(["Конверсия в корзину, %", 10.0, 7.0, -3.0])
    ws.append(["Конверсия в заказ, %", 2.0, 5.0, 3.0])
    buf = BytesIO()
    wb.save(buf)

    out = parse_wb_competitor_excel(content=buf.getvalue(), report_date=date(2026, 5, 11), period="week")
    items = out["items"]
    # our values
    ctr_111 = next(i for i in items if i["nm_id"] == 111 and i["metric_code"] == "ctr")
    assert ctr_111["our_value"] == 5.0
    assert ctr_111["competitor_median_value"] == 15.0
    traffic_222 = next(i for i in items if i["nm_id"] == 222 and i["metric_code"] == "traffic")
    assert traffic_222["our_value"] == 300.0
    assert traffic_222["competitor_median_value"] == 100.0  # поле по смыслу: среднее по конкурентам для «Показы»
    assert traffic_222.get("extra", {}).get("competitor_aggregate") == "mean"

    # Конверсии — только из строк «Конверсия …» в Excel (п.п.; строки «…, шт» в импорт конверсий не идут).
    cart_111 = next(i for i in items if i["nm_id"] == 111 and i["metric_code"] == "funnel_cart")
    assert cart_111["our_value"] == 10.0
    assert cart_111["competitor_median_value"] == 7.0
    assert cart_111["unit"] == "%"
    cart_222 = next(i for i in items if i["nm_id"] == 222 and i["metric_code"] == "funnel_cart")
    assert cart_222["our_value"] == 7.0
    assert cart_222["competitor_median_value"] == 10.0
    ord_111 = next(i for i in items if i["nm_id"] == 111 and i["metric_code"] == "funnel_order")
    assert ord_111["our_value"] == 2.0
    assert ord_111["competitor_median_value"] == 5.0
    ord_222 = next(i for i in items if i["nm_id"] == 222 and i["metric_code"] == "funnel_order")
    assert ord_222["our_value"] == 5.0
    assert ord_222["competitor_median_value"] == 2.0


def test_parse_pokazateli_conversion_strict_row_labels_import_values_as_given() -> None:
    """Строго строка «Конверсия в корзину, %» — импортируем числа как в Excel (контроль качества — в аналитике)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Показатели"
    ws.append(["x"])
    ws.append(["Показатели", "Артикул WB 10", "Артикул WB 20"])
    ws.append(["Показы", 1000, 2000])
    ws.append(["CTR", 3.0, 4.0])
    ws.append(["Конверсия в корзину, %", 40.0, 200.0])
    ws.append(["Конверсия в заказ, %", 1.0, 2.0])
    buf = BytesIO()
    wb.save(buf)

    out = parse_wb_competitor_excel(content=buf.getvalue(), report_date=date(2026, 5, 11), period="week")
    cart = next(i for i in out["items"] if i["nm_id"] == 10 and i["metric_code"] == "funnel_cart")
    assert cart["our_value"] == 40.0
    assert cart["competitor_median_value"] == 200.0
    assert cart.get("extra", {}).get("competitor_aggregate") == "median"


def test_parse_pokazateli_ctr_fraction_cells_become_percent_points() -> None:
    """CTR: значения строго между 0 и 1 считаем долей и умножаем на 100 (п.п.)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Показатели"
    ws.append(["x"])
    ws.append(["Показатели", "Артикул WB 10", "Артикул WB 20"])
    ws.append(["Показы", 1000, 2000])
    ws.append(["CTR", 0.031, 0.04])
    ws.append(["Конверсия в корзину, %", 8.0, 12.0])
    ws.append(["Конверсия в заказ, %", 2.0, 3.0])
    buf = BytesIO()
    wb.save(buf)

    out = parse_wb_competitor_excel(content=buf.getvalue(), report_date=date(2026, 5, 11), period="week")
    ctr10 = next(i for i in out["items"] if i["nm_id"] == 10 and i["metric_code"] == "ctr")
    assert ctr10["our_value"] == pytest.approx(3.1)
    assert ctr10["competitor_median_value"] == pytest.approx(4.0)


def test_parse_pokazateli_traffic_competitor_aggregate_is_mean() -> None:
    """«Показы»: по конкурентам — среднее (четыре артикула)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Показатели"
    ws.append(["x"])
    ws.append(["Показатели", "Артикул WB 1", "Артикул WB 2", "Артикул WB 3", "Артикул WB 4"])
    ws.append(["Показы", 100, 300, 500, 900])
    ws.append(["CTR", 1.0, 2.0, 3.0, 4.0])
    ws.append(["Конверсия в корзину, %", 1.0, 2.0, 3.0, 4.0])
    ws.append(["Конверсия в заказ, %", 1.0, 2.0, 3.0, 4.0])
    buf = BytesIO()
    wb.save(buf)

    out = parse_wb_competitor_excel(content=buf.getvalue(), report_date=date(2026, 5, 11), period="week")
    t1 = next(i for i in out["items"] if i["nm_id"] == 1 and i["metric_code"] == "traffic")
    assert t1["our_value"] == 100.0
    assert t1["competitor_median_value"] == pytest.approx((300 + 500 + 900) / 3)
    assert t1.get("extra", {}).get("competitor_aggregate") == "mean"


def test_parse_pokazateli_median_excludes_zero_competitor_values() -> None:
    """Нули у конкурентов не входят в медиану (остальные товары)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Показатели"
    ws.append(["x"])
    ws.append(["Показатели", "Артикул WB 10", "Артикул WB 11", "Артикул WB 12", "Разница"])
    ws.append(["Показы", 100, 200, 150, 0])
    ws.append(["CTR", 5.0, 6.0, 4.0, 0])
    ws.append(["Конверсия в корзину, %", 10.0, 0.0, 8.0, 0])
    ws.append(["Конверсия в заказ, %", 1.0, 1.0, 1.0, 0])
    buf = BytesIO()
    wb.save(buf)

    out = parse_wb_competitor_excel(content=buf.getvalue(), report_date=date(2026, 5, 11), period="week")
    c10 = next(i for i in out["items"] if i["nm_id"] == 10 and i["metric_code"] == "funnel_cart")
    assert c10["our_value"] == 10.0
    assert c10["competitor_median_value"] == 8.0
    c11 = next(i for i in out["items"] if i["nm_id"] == 11 and i["metric_code"] == "funnel_cart")
    assert c11["our_value"] == 0.0
    assert c11["competitor_median_value"] == pytest.approx(9.0)
    c12 = next(i for i in out["items"] if i["nm_id"] == 12 and i["metric_code"] == "funnel_cart")
    assert c12["our_value"] == 8.0
    assert c12["competitor_median_value"] == 10.0


def test_parse_pokazateli_funnel_median_after_mixed_fraction_and_percent_points() -> None:
    """Смесь п.п. (15) и долей Excel (0.12) в одной строке — медиана по конкурентам в п.п., не по сырым 0.12."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Показатели"
    ws.append(["x"])
    ws.append(["Показатели", "Артикул WB 10", "Артикул WB 20", "Артикул WB 30"])
    ws.append(["Показы", 1000, 2000, 3000])
    ws.append(["CTR", 3.0, 4.0, 5.0])
    ws.append(["Конверсия в корзину, %", 8.0, 12.0, 10.0])
    ws.append(["Конверсия в заказ, %", 15.0, 0.12, 0.18])
    buf = BytesIO()
    wb.save(buf)

    out = parse_wb_competitor_excel(content=buf.getvalue(), report_date=date(2026, 5, 11), period="week")
    o10 = next(i for i in out["items"] if i["nm_id"] == 10 and i["metric_code"] == "funnel_order")
    assert o10["our_value"] == 15.0
    assert o10["competitor_median_value"] == pytest.approx(15.0)


def test_parse_pokazateli_funnel_all_open_fractions_become_percent_points() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Показатели"
    ws.append(["x"])
    ws.append(["Показатели", "Артикул WB 1", "Артикул WB 2", "Артикул WB 3"])
    ws.append(["Показы", 100, 200, 300])
    ws.append(["CTR", 2.0, 3.0, 4.0])
    ws.append(["Конверсия в корзину, %", 0.08, 0.10, 0.12])
    ws.append(["Конверсия в заказ, %", 0.02, 0.03, 0.04])
    buf = BytesIO()
    wb.save(buf)

    out = parse_wb_competitor_excel(content=buf.getvalue(), report_date=date(2026, 5, 11), period="week")
    fo1 = next(i for i in out["items"] if i["nm_id"] == 1 and i["metric_code"] == "funnel_order")
    assert fo1["our_value"] == pytest.approx(2.0)
    assert fo1["competitor_median_value"] == pytest.approx(3.5)

